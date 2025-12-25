from openpyxl import load_workbook

# from ..normalizer import normalize_label
import re


def normalize_label(label: str) -> str:
    if not label:
        return ""
    label = label.lower().strip()
    label = re.sub(r"[^\w\s]", "", label)
    label = re.sub(r"\s+", "_", label)
    return label


def infer_currency(label: str) -> str | None:
    if "£" in label or "gbp" in label.lower():
        return "GBP"
    if "$" in label or "usd" in label.lower():
        return "USD"
    if "€" in label or "eur" in label.lower():
        return "EUR"
    return None


def infer_column_type(ws, col_idx: int, start_row: int, sample_size: int = 10) -> str:
    numeric = 0
    text = 0

    for r in range(start_row, start_row + sample_size):
        val = ws.cell(row=r, column=col_idx).value
        if isinstance(val, (int, float)):
            numeric += 1
        elif isinstance(val, str):
            text += 1

    if numeric > text:
        return "number"
    return "string"


def detect_header_row(ws, max_scan_rows: int = 20) -> int | None:
    for row_idx in range(1, max_scan_rows + 1):
        values = [cell.value for cell in ws[row_idx] if isinstance(cell.value, str)]

        if len(values) >= 2:
            return row_idx

    return None


def detect_columns(ws, header_row: int) -> list[dict]:
    columns = []

    for idx, cell in enumerate(ws[header_row]):
        if not cell.value:
            continue

        label = str(cell.value).strip()
        normalized = normalize_label(label)

        column_type = infer_column_type(ws, idx + 1, header_row + 1)

        col = {
            "index": idx,
            "label": label,
            "normalized": normalized,
            "type": column_type,
        }

        if column_type == "currency":
            col["currency"] = infer_currency(label)

        columns.append(col)

    return columns


def detect_totals(ws, columns: list[dict]) -> dict:
    total_columns = []
    total_rows = []

    for col in columns:
        if "total" in col["normalized"]:
            total_columns.append(col["index"])

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "total" in cell.value.lower():
                total_rows.append(cell.row)

    return {
        "row_indices": total_rows,
        "columns": total_columns,
    }


def detect_sections(columns: list[dict]) -> dict | None:
    for col in columns:
        if col["normalized"] in ("category", "budget_category", "section"):
            return {"type": "grouped_rows", "group_by_column": col["index"]}
    return None


def detect_excel_structure(file_path: str) -> dict:
    wb = load_workbook(file_path, data_only=False)

    structure = {"version": 1, "sheets": []}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        header_row = detect_header_row(ws)
        if header_row is None:
            continue

        columns = detect_columns(ws, header_row)
        totals = detect_totals(ws, columns)

        structure["sheets"].append(
            {
                "sheet_name": sheet_name,
                "is_primary": sheet_name.lower() in ("budget", "summary"),
                "header": {
                    "row_index": header_row,
                    "raw_labels": [c["label"] for c in columns],
                },
                "data": {
                    "start_row": header_row + 1,
                    "end_row": None,
                },
                "columns": columns,
                "sections": detect_sections(columns),
                "totals": totals,
            }
        )

    return structure


def classify_row(text: str) -> str | None:
    t = text.lower().strip()

    if not t:
        return None

    if re.match(r"^\d+\.\s+", t):
        return "category"

    if t.startswith("total ") and "project" not in t:
        return "category_total"

    if "total project" in t:
        return "grand_total"

    if any(k in t for k in ["organisation", "project name", "project period"]):
        return "metadata"

    if any(k in t for k in ["authorised", "signatory", "contact person"]):
        return "signature"

    return "example_item"


def detect_row_semantic_structure(file_path: str) -> dict:
    wb = load_workbook(file_path)
    ws = wb.active

    rows = []
    current_category = None

    for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first_cell = row[0] if row else None
        if not isinstance(first_cell, str):
            continue

        text = first_cell.strip()
        row_type = classify_row(text)
        if not row_type:
            continue

        entry = {
            "row": r_idx,
            "type": row_type,
            "label": text,
        }

        if row_type == "category":
            match = re.match(r"^(\d+)\.\s+(.*)", text)
            entry["category_index"] = match.group(1)
            entry["label"] = match.group(2)
            current_category = entry["label"]

        elif row_type == "example_item":
            entry["belongs_to_category"] = current_category

        rows.append(entry)

    return {
        "version": 2,
        "sheet": ws.title,
        "rows": rows,
    }


if __name__ == "__main__":

    import json

    # if len(sys.argv) != 2:
    #     print("Usage: python detector.py <excel_file_path>")
    #     sys.exit(1)

    # file_path = sys.argv[1]

    file_path = "/home/noro/repos/GrandFlow/uploads/budget/Donor_budget_template.xlsx"
    # detected_structure = detect_excel_structure(file_path)
    detected_structure = detect_row_semantic_structure(file_path)
    print(json.dumps(detected_structure, indent=2))

"""Spreadsheet structure detection utilities.

Public API
- ExcelStructureDetector(file_path)
    - read_sheet_with_pandas() -> DataFrame
    - read_sheet_with_openpyxl() -> (DataFrame, formula_flags)
    - filter_out_formula_rows(df) -> DataFrame
    - normalize_dataframe(df) -> DataFrame
    - remove_numeric_rows(df) -> DataFrame
    - to_detection_json(df) -> str
    - detect_structure() -> DataFrame  # high-level pipeline

These helpers read Excel sheets and return a cleaned DataFrame where
rows that are formulas or mostly numeric are removed, ready for
downstream template detection logic.
"""

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# from ..normalizer import normalize_label
import re
import json
import pandas as pd
import numpy as np


# numeric_pattern = re.compile(r"^\s*[-+]?\d*\.?\d+\s*$")
numeric_pattern = re.compile(
    r"""
    ^\s*
    (
        [-+]?\d+(\.\d*)?      # allows 1, 1., 1.5, 2.0
        |
        [-+]?\d+\s*-\s*\d+    # numeric range
    )
    \s*$
    """,
    re.VERBOSE,
)


def is_numeric(val):
    """Return True if value is numeric-like or empty/None.

    This is a permissive check used to filter out rows that are
    primarily numeric (e.g. amounts) when detecting textual labels.
    """

    if val is None:
        return True
    val_str = str(val).strip()
    if val_str == "":
        return True
    return bool(numeric_pattern.match(val_str))


class ExcelStructureDetector:
    def __init__(self, file_path: str):

        self.file_path = file_path
        self.data = []
        self.formula_flags = []
        self.wb = load_workbook(file_path, data_only=False)
        self.ws = self.wb.active

    def read_sheet_with_pandas(self) -> pd.DataFrame:
        """Read the active sheet into a pandas DataFrame (text only).

        Keeps everything as strings and does not treat empty cells as NaN.
        """
        return pd.read_excel(
            self.file_path,
            header=None,  # ❗ critical
            dtype=str,  # keep everything as text
            keep_default_na=False,
        )

    def read_sheet_with_openpyxl(self) -> tuple[pd.DataFrame, list[list[bool]]]:
        """Read the sheet using openpyxl and also return formula flags.

        Returns a tuple (DataFrame, formula_flags) where formula_flags is
        a list of rows containing booleans for whether each cell had a formula.
        """
        first_row = [""] + [get_column_letter(i + 1) for i in range(len(self.ws[1]))]
        self.data = [first_row]
        self.formula_flags = [first_row]
        for i, row in enumerate(self.ws.iter_rows(values_only=False)):
            row_values = [i + 1]
            row_formulas = [i + 1]
            for j, cell in enumerate(row):

                # If cell has a formula, store it separately
                if cell.data_type == "f":  # 'f' indicates formula in openpyxl
                    row_values.append(cell.value)  # You can also keep calculated value instead
                    row_formulas.append(True)  # Mark as formula
                else:
                    row_values.append(cell.value)
                    row_formulas.append(False)  # Not a formula
            self.data.append(row_values)
            self.formula_flags.append(row_formulas)

        return pd.DataFrame(self.data), self.formula_flags

    def filter_out_formula_rows(self) -> pd.DataFrame:
        """
        Keep only rows where no cell (except first column) has a formula.
        """
        # Convert formula_flags to DataFrame for easy row-wise operations
        ff_df = pd.DataFrame(self.formula_flags)
        df = pd.DataFrame(self.data)
        # We assume first column is Excel row numbers, so skip it
        # if first column of formula_flags corresponds to row numbers, adjust if needed
        ff_df_body = ff_df.iloc[1:, 1:]

        # Keep rows where **no cell has formula**
        mask = ~ff_df_body.any(axis=1)  # True for rows without formulas

        # Apply mask to original df (skip header row if you have one)
        filtered_df = df.iloc[1:][mask].reset_index(drop=True)

        # Optionally, keep the header row at top
        header_row = df.iloc[0:1]
        filtered_df = pd.concat([header_row, filtered_df]).reset_index(drop=True)

        return filtered_df

    def normalize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove empty rows/columns except header/row number column"""
        header = df.iloc[0]
        body = df.iloc[1:]

        # drop completely empty rows (ignore ExcelRow column)
        body = body.dropna(how="all", subset=body.columns[1:])
        return pd.concat([header.to_frame().T, body]).reset_index(drop=True)
        # return df.replace(r"^\s*$", None, regex=True).dropna(how="all").reset_index(drop=True)

    def remove_numeric_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove rows that are entirely numeric (except first column and row)."""

        # first row is header (A, B, C...) and first column is row numbers / numeration
        header = df.iloc[0:1]
        body = df.iloc[1:]

        # Create a boolean mask: True for cells that are numeric-like
        mask_numeric = body.iloc[:, 1:].applymap(
            lambda v: bool(v) and bool(numeric_pattern.match(str(v))) if v is not None else True
        )

        # Replace numeric-like values with None
        body.iloc[:, 1:] = body.iloc[:, 1:].mask(mask_numeric, other=None)

        # Drop rows that are all NaN except the first column
        body_cleaned = body.dropna(how="all", subset=body.columns[1:]).reset_index(drop=True)

        # Combine with header
        df_cleaned = pd.concat([header, body_cleaned]).reset_index(drop=True)
        return df_cleaned

    def filter_list_of_possible_fields(self, df: pd.DataFrame) -> list[str]:
        """Extract a list of possible field names from the cleaned DataFrame."""
        possible_fields = set()

        for r_idx, row in df.iloc[1:].iterrows():  # skip header
            for c_idx, val in enumerate(row[1:], start=1):  # skip first column (row numbers)
                if val and isinstance(val, str) and not is_numeric(val):
                    possible_fields.add(val.strip())

        return list(possible_fields)

    def to_detection_json(self, df: pd.DataFrame) -> list[dict]:
        """Serialize detection results (from cleaned DataFrame) to a JSON string."""

        output = []

        for r_idx, row in df.iloc[1:].iterrows():  # skip header
            for c_idx, val in enumerate(row[1:], start=1):  # skip first column (row numbers)
                if not is_numeric(val):
                    coordinate = f"{df.iloc[0, c_idx]}{row[0]}"  # column letter + row number
                    # For now, dummy suggested field + confidence
                    suggested_field = "unknown"
                    confidence = 0.0
                    output.append(
                        {
                            "coordinate": coordinate,
                            "row": row[0],
                            "col": df.iloc[0, c_idx],
                            "value": val,
                            "suggested_field": suggested_field,
                            "confidence": confidence,
                        }
                    )

        # Dump to JSON
        # json_str = json.dumps(output, indent=2)
        return output

    def detect_structure(self) -> pd.DataFrame:
        """High-level pipeline: read with openpyxl, filter formulas, normalize, remove numeric rows.

        Returns the cleaned DataFrame ready for detection/serialization.
        """
        df, _ = self.read_sheet_with_openpyxl()
        df = self.filter_out_formula_rows()
        df = self.normalize_dataframe(df)
        df = self.remove_numeric_rows(df)
        return df


if __name__ == "__main__":

    import json

    # if len(sys.argv) != 2:
    #     print("Usage: python detector.py <excel_file_path>")
    #     sys.exit(1)

    # file_path = sys.argv[1]

    file_path = "/home/noro/repos/GrandFlow/uploads/budget/Donor_budget_template.xlsx"
    # detected_structure = detect_excel_structure(file_path)
    # df = load_raw_sheet(file_path)
    excel_reader = ExcelStructureDetector(file_path)
    # Use the high-level pipeline to get a cleaned DataFrame
    df = excel_reader.detect_structure()

    # Serialize detections to JSON
    json_output = excel_reader.to_detection_json(df)
    print(json_output)

    # detected_structure = detect_structure(df)
    # print(json.dumps(detected_structure, indent=2))
